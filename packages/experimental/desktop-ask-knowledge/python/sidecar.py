#!/usr/bin/env python3
"""JSON stdin/stdout entry for the octopus_DSH ask-knowledge sidecar."""

from __future__ import annotations

import hashlib
import json
import os
import sys
import tempfile
import traceback
from pathlib import Path


def _kb_root() -> Path:
    env = os.environ.get("OCTOPUS_KB_ROOT", "").strip()
    if env:
        return Path(env)
    here = Path(__file__).resolve().parent
    vendored = here / "kb"
    if (vendored / "prompts" / "propose.md").is_file():
        return vendored
    meipass = getattr(sys, "_MEIPASS", None)
    if meipass is not None:
        return Path(meipass)
    return here


def _ensure_path() -> Path:
    root = _kb_root()
    src = root / "src"
    if src.is_dir() and str(src) not in sys.path:
        sys.path.insert(0, str(src))
    os.environ.setdefault("OCTOPUS_KB_ROOT", str(root))
    return root


def _read_page_text(vault: Path, rel_path: str, limit: int = 4000) -> str:
    path = vault / rel_path
    if not path.is_file():
        return ""
    text = path.read_text(encoding="utf-8", errors="replace")
    return text if len(text) <= limit else text[:limit]


def _kind_for(path: str, bundle: dict) -> str:
    concepts = {item.get("path") for item in bundle.get("concepts", [])}
    entities = {item.get("path") for item in bundle.get("entities", [])}
    raws = {item.get("path") for item in bundle.get("raw_sources", [])}
    if path in concepts:
        return "concept"
    if path in entities:
        return "entity"
    if path in raws:
        return "raw"
    if path.endswith("AGENTS.md"):
        return "concept"
    return "raw"


def cmd_self_test(req: dict) -> dict:
    root = _ensure_path()
    import octopus_kb_compound  # noqa: F401

    prompt = root / "prompts" / "propose.md"
    schema = root / "schemas" / "llm" / "proposal.json"
    if not prompt.is_file():
        raise RuntimeError(f"missing {prompt}")
    if not schema.is_file():
        raise RuntimeError(f"missing {schema}")
    from octopus_kb_compound.proposals import _schema_path, kb_resource_root
    from octopus_kb_compound.apply import _builtins_rules_path
    from octopus_kb_compound.propose import fill_create_page_frontmatter

    proposal_schema = _schema_path()
    if not proposal_schema.is_file():
        raise RuntimeError(f"missing proposal schema {proposal_schema}")
    rules = _builtins_rules_path()
    if not rules.is_file():
        raise RuntimeError(f"missing apply rules {rules}")
    rules_schema = kb_resource_root() / "schemas" / "rules" / "v1.json"
    if not rules_schema.is_file():
        raise RuntimeError(f"missing rules schema {rules_schema}")
    try:
        import markitdown  # noqa: F401

        has_markitdown = True
    except ImportError:
        has_markitdown = False
    try:
        import pdfminer  # noqa: F401
        import pdfplumber  # noqa: F401

        has_pdf = True
    except ImportError:
        has_pdf = False
    try:
        import openpyxl  # noqa: F401

        has_xlsx = True
    except ImportError:
        has_xlsx = False
    filled = {
        "operations": [
            {
                "op": "create_page",
                "frontmatter": {"title": "t", "type": "concept", "lang": "zh"},
                "body": "b",
            }
        ]
    }
    fill_create_page_frontmatter(filled)
    frontmatter = filled["operations"][0]["frontmatter"]
    if frontmatter.get("role") != "concept" or frontmatter.get("layer") != "wiki":
        raise RuntimeError("fill_create_page_frontmatter did not set role and layer")
    _assert_retrieve_fold()
    _assert_convert_file()
    key = os.environ.get("DEEPSEEK_API_KEY", "")
    result = {
        "ok": True,
        "compound": True,
        "prompts": True,
        "schemas": True,
        "proposalSchema": True,
        "applyRules": True,
        "rulesSchema": True,
        "pageMetaFill": True,
        "retrieveFold": True,
        "convertFile": True,
        "markitdown": has_markitdown,
        "pdf": has_pdf,
        "xlsx": has_xlsx,
        "hasDeepseekKey": bool(key),
    }
    if req.get("expectKey") is True:
        result["hasDeepseekKey"] = bool(key)
    if os.environ.get("ASK_KNOWLEDGE_SIDECAR_KEY_HASH") == "1":
        result["keyHash"] = hashlib.sha256(key.encode("utf-8")).hexdigest()
    return result


def _assert_retrieve_fold() -> None:
    from octopus_kb_compound.retrieve import build_retrieval_bundle

    # ⾯向…“⼀⽼⼀⼩” — Kangxi radicals that NFKC-fold to 一老一小
    body = "\u2faf\u5411\u2fbc\u8d28\u91cf\u53d1\u5c55\u7684\u201c\u2f00\u2f7c\u2f00\u2f29\u201d\u653f\u7b56"
    vault = Path(tempfile.mkdtemp(prefix="kb-retrieve-fold-"))
    page = vault / "raw" / "kangxi.md"
    page.parent.mkdir(parents=True, exist_ok=True)
    page.write_text(
        "---\ntitle: policy\nrole: raw_source\n---\n" + body + "\n",
        encoding="utf-8",
    )
    bundle = build_retrieval_bundle(vault, "一老一小")
    if not bundle.raw_source_items:
        raise RuntimeError("retrieve did not fold Kangxi radicals to CJK")


def _assert_convert_file() -> None:
    folder = Path(tempfile.mkdtemp(prefix="kb-convert-"))
    source = folder / "note.md"
    source.write_text("# 仅本会话\n正文\n", encoding="utf-8")
    result = cmd_convert_file({"path": str(source)})
    if not result.get("ok") or "正文" not in result.get("body", ""):
        raise RuntimeError("convert-file did not return markdown body")
    if (folder / "raw").is_dir():
        raise RuntimeError("convert-file wrote a vault")


def cmd_bootstrap(req: dict) -> dict:
    vault = Path(req["vault"])
    vault.mkdir(parents=True, exist_ok=True)
    (vault / "wiki").mkdir(exist_ok=True)
    (vault / "raw").mkdir(exist_ok=True)
    (vault / ".octopus-kb").mkdir(exist_ok=True)
    config = vault / ".octopus-kb" / "config.toml"
    if not config.is_file():
        config.write_text(
            "version = 1\n"
            "[llm]\n"
            'default_profile = "deepseek"\n'
            "[llm.profiles.deepseek]\n"
            'base_url = "https://api.deepseek.com"\n'
            'model = "deepseek-v4-flash"\n'
            'api_key_env = "DEEPSEEK_API_KEY"\n',
            encoding="utf-8",
        )
    return {"ok": True, "vault": str(vault)}


def cmd_recover(req: dict) -> dict:
    _ensure_path()
    from octopus_kb_compound.apply import recover_proposal

    result = recover_proposal(req["proposalId"], req["vault"]).to_dict()
    return {"ok": True, **result}


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("|", "\\|").replace("\n", " ")


def _markdown_table(rows: list[list[str]]) -> str:
    if not rows:
        return ""
    width = max(len(row) for row in rows)
    padded = [row + [""] * (width - len(row)) for row in rows]
    header = padded[0]
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(["---"] * width) + " |",
    ]
    for row in padded[1:]:
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines)


def _decode_text_file(source: Path) -> str:
    raw = source.read_bytes()
    if raw.startswith(b"\xef\xbb\xbf"):
        return raw.decode("utf-8-sig")
    try:
        from charset_normalizer import from_bytes

        best = from_bytes(raw).best()
        if best is not None:
            return str(best)
    except ImportError:
        pass
    for encoding in ("utf-8", "gb18030", "gbk"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _csv_to_markdown(source: Path) -> tuple[str, dict[str, str]]:
    import csv
    import io

    rows = [[_cell_text(cell) for cell in row] for row in csv.reader(io.StringIO(_decode_text_file(source)))]
    return _markdown_table(rows), {
        "title": source.stem,
        "source_file": source.name,
        "original_format": "csv",
        "ingest_method": "csv",
    }


def _xlsx_to_markdown(source: Path) -> tuple[str, dict[str, str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(source, read_only=True, data_only=True)
    parts: list[str] = []
    try:
        for name in workbook.sheetnames:
            sheet = workbook[name]
            rows = [[_cell_text(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
            rows = [row for row in rows if any(cell.strip() for cell in row)]
            if not rows:
                continue
            parts.append(f"## {name}\n\n{_markdown_table(rows)}")
    finally:
        workbook.close()
    body = "\n\n".join(parts)
    return body, {
        "title": source.stem,
        "source_file": source.name,
        "original_format": "xlsx",
        "ingest_method": "openpyxl",
    }


def _convert_source(source: Path, *, session_attachment: bool = False) -> dict:
    from octopus_kb_compound.ingest import OptionalDependencyMissing, convert_file_to_markdown

    ext = source.suffix.lower()
    if ext in {".md", ".txt"}:
        body = source.read_text(encoding="utf-8", errors="replace")
        metadata = {
            "title": source.stem,
            "source_file": source.name,
            "original_format": ext.lstrip("."),
            "ingest_method": "copy",
        }
    elif ext == ".xlsx":
        try:
            body, metadata = _xlsx_to_markdown(source)
        except ImportError as exc:
            return {"ok": False, "error": str(exc), "code": "type-unsupported"}
        except Exception as exc:  # noqa: BLE001 — sidecar ingest boundary
            return {"ok": False, "error": str(exc), "code": "ingest-failed"}
    elif ext == ".csv":
        try:
            body, metadata = _csv_to_markdown(source)
        except Exception as exc:  # noqa: BLE001 — sidecar ingest boundary
            return {"ok": False, "error": str(exc), "code": "ingest-failed"}
    else:
        try:
            body, metadata = convert_file_to_markdown(str(source))
        except OptionalDependencyMissing as exc:
            return {"ok": False, "error": str(exc), "code": "type-unsupported"}
        except Exception as exc:  # noqa: BLE001 — sidecar ingest boundary
            return {"ok": False, "error": str(exc), "code": "ingest-failed"}
    if body.strip() == "":
        if ext == ".pdf":
            empty = (
                "这份 PDF 没有可提取的文字。扫描件还不能作为会话附件。"
                if session_attachment
                else "这份 PDF 没有可提取的文字。扫描件还不能入库。"
            )
        elif ext in {".csv", ".xlsx"}:
            empty = "这份表格是空的。"
        else:
            empty = "这份文件没有可提取的文字。"
        return {"ok": False, "error": empty, "code": "ingest-failed"}
    return {"ok": True, "body": body, "metadata": metadata}


def cmd_ingest_file(req: dict) -> dict:
    _ensure_path()
    from octopus_kb_compound.ingest import generate_raw_page

    source = Path(req["path"])
    vault = Path(req["vault"])
    converted = _convert_source(source)
    if not converted.get("ok"):
        return converted
    dest = generate_raw_page(converted["body"], converted["metadata"], vault / "raw", lang="zh")
    rel = dest.relative_to(vault).as_posix()
    return {"ok": True, "rawRelPath": rel}


def cmd_convert_file(req: dict) -> dict:
    _ensure_path()
    source = Path(req["path"])
    ext = source.suffix.lower()
    if ext not in {".md", ".txt", ".html", ".htm", ".pdf"}:
        return {
            "ok": False,
            "error": "仅支持 Markdown、TXT、HTML、PDF。表格请走问数。",
            "code": "type-unsupported",
        }
    converted = _convert_source(source, session_attachment=True)
    if not converted.get("ok"):
        return converted
    metadata = converted["metadata"]
    return {
        "ok": True,
        "body": converted["body"],
        "title": metadata["title"],
        "sourceFile": metadata["source_file"],
    }


def cmd_propose(req: dict) -> dict:
    _ensure_path()
    from octopus_kb_compound.propose import ProposeInputError, ProposeRuntimeError, propose_from_raw

    try:
        result = propose_from_raw(req["rawFile"], req["vault"]).to_dict()
    except (ProposeInputError, ProposeRuntimeError) as exc:
        return {"ok": False, "error": str(exc)}
    return {"ok": True, **result}


def cmd_validate_apply(req: dict) -> dict:
    _ensure_path()
    from octopus_kb_compound.apply import ValidateInputError, ValidateRuntimeError, validate_proposal_file

    try:
        result = validate_proposal_file(req["proposal"], req["vault"], apply=True).to_dict()
    except (ValidateInputError, ValidateRuntimeError) as exc:
        return {"ok": False, "error": str(exc)}
    return {"ok": True, **result}


def cmd_retrieve_bundle(req: dict) -> dict:
    _ensure_path()
    from octopus_kb_compound import retrieve as retrieve_mod
    from octopus_kb_compound.retrieve import build_retrieval_bundle

    vault = Path(req["vault"])
    query = req["query"]
    from octopus_kb_compound.heal import heal_page_meta_rejections

    heal_page_meta_rejections(vault)
    bundle = build_retrieval_bundle(vault, query)
    try:
        retrieve_mod._touch_marker(vault)
    except OSError:
        pass
    data = bundle.to_dict()
    inner = data.get("bundle", {})
    items = []
    seen: set[str] = set()
    for group, kind in (
        (inner.get("concepts", []), "concept"),
        (inner.get("entities", []), "entity"),
        (inner.get("raw_sources", []), "raw"),
    ):
        for item in group:
            path = item.get("path")
            if not path or path in seen:
                continue
            seen.add(path)
            items.append(
                {
                    "path": path,
                    "title": item.get("title") or path,
                    "reason": item.get("reason") or "",
                    "kind": kind,
                    "text": _read_page_text(vault, path),
                }
            )
    return {
        "ok": True,
        "query": query,
        "items": items,
        "warnings": data.get("warnings", []),
        "token_estimate": data.get("token_estimate", 0),
    }


def cmd_lookup(req: dict) -> dict:
    _ensure_path()
    from octopus_kb_compound.lookup import lookup_term

    vault = Path(req["vault"])
    term = req["term"]
    from octopus_kb_compound.heal import heal_page_meta_rejections

    heal_page_meta_rejections(vault)
    result = lookup_term(term, vault).to_dict()
    canonical = result.get("canonical")
    text = ""
    path = None
    if isinstance(canonical, dict):
        path = canonical.get("path")
        if path:
            text = _read_page_text(vault, path)
    return {
        "ok": True,
        "term": term,
        "canonicalPath": path,
        "text": text or None,
        "ambiguous": result.get("ambiguous", False),
        "collisions": result.get("collisions", []),
    }


def cmd_inbox_list(req: dict) -> dict:
    _ensure_path()
    from octopus_kb_compound.inbox import list_inbox

    result = list_inbox(req["vault"])
    if isinstance(result, dict):
        items = result.get("items") or result.get("deferred") or []
        return {"ok": True, "deferredCount": len(items), "items": items}
    if isinstance(result, list):
        return {"ok": True, "deferredCount": len(result), "items": result}
    return {"ok": True, "deferredCount": 0, "items": []}


COMMANDS = {
    "self-test": cmd_self_test,
    "bootstrap": cmd_bootstrap,
    "recover": cmd_recover,
    "ingest-file": cmd_ingest_file,
    "convert-file": cmd_convert_file,
    "propose": cmd_propose,
    "validate-apply": cmd_validate_apply,
    "retrieve-bundle": cmd_retrieve_bundle,
    "lookup": cmd_lookup,
    "inbox-list": cmd_inbox_list,
}


def main() -> int:
    raw = sys.stdin.read()
    try:
        req = json.loads(raw) if raw.strip() else {}
    except json.JSONDecodeError as exc:
        json.dump({"ok": False, "error": f"invalid json: {exc}"}, sys.stdout, ensure_ascii=False)
        sys.stdout.write("\n")
        return 2
    command = req.get("command") or req.get("cmd")
    handler = COMMANDS.get(command)
    if handler is None:
        json.dump({"ok": False, "error": f"unknown command: {command}"}, sys.stdout, ensure_ascii=False)
        sys.stdout.write("\n")
        return 2
    try:
        result = handler(req)
    except Exception as exc:  # noqa: BLE001 — sidecar boundary
        traceback.print_exc(file=sys.stderr)
        json.dump({"ok": False, "error": str(exc)}, sys.stdout, ensure_ascii=False)
        sys.stdout.write("\n")
        return 1
    if "ok" not in result:
        result = {"ok": True, **result}
    json.dump(result, sys.stdout, ensure_ascii=False)
    sys.stdout.write("\n")
    return 0 if result.get("ok") else 1


if __name__ == "__main__":
    raise SystemExit(main())
